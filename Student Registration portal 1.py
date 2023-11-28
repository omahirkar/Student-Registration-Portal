import tkinter
from tkinter import*
from datetime import date
from tkinter import filedialog
from tkinter import messagebox
from PIL import Image,ImageTk
import os
from tkinter.ttk import Combobox
import openpyxl , xlrd
from openpyxl import Workbook 
import pathlib

background= "#06283D"
framebg="#EDEDED"
framefg="#06283D"

root = Tk()
root.title("Student Registration System")
root.geometry("1250x750+210+100")
root.config(bg=background)


file=pathlib.Path("Student_data.xlsx")
if file.exists():
    pass
else:
    file=Workbook()
    sheet=file.active
    sheet['A1']="PRN no"
    sheet['B1']="Name"
    sheet['C1']="Year"
    sheet['D1']="Gender"
    sheet['E1']="Course"
    sheet['F1']="Date of Registration"
    sheet['G1']="Mobile number"
    sheet['H1']="Email"
    sheet['I1']="Date Of Birth"
    sheet['J1']="City"
    sheet['K1']="Mother Name"
    sheet['L1']="Father Name"

    file.save("Student_info_data.xlsx")

############################Exit Window#########################################
def Exit():
    root.destroy()

##############################Show_Image###########################################
def showimage():
    global filename
    global img
    filename=filedialog.askopenfilename(initialdir=os.getcwd(),
                                        title="Select image file",filetypes=(("JPG File","*.jpg"),
                                                                             ("PNG File","*.png"),
                                                                             ("ALL files","*.txt")))
    img = (Image.open(filename))
    resized_image=img.resize((190,190))
    photo2 = Image.PhotoImage(resized_image)
    lbl.config(image=photo2)
    lbl.image=photo2

###########################PRN No###############################################
def PRN_no():
    file=openpyxl.load_workbook('Student_info_data.xlsx')
    sheet=file.active
    row=sheet.max_row

    max_row_value=sheet.cell(row=row,column=203033124601).value

    try:
        PRN.set(max_row_value+1)
    except:
        PRN.set("203033124601")

#####################################Clear#########################################
def Clear():
    global img
    Name.set('')
    DOB_Name.set('')
    Year.set('Select Year')
    Course.set('')
    Father_Name.set('')
    Mother_Name.set('')
    City_Name.set('')
    Email.set('')

    PRN_no()

    saveButton.config(state='normal')

    img1=PhotoImage(file="F:\\New folder (2)\\Documents\\Python\\upload photo.png")
    lbl.config(image=img1)
    lbl.image=img1

    img=""
#####################################Save#########################################
def Save():
    R1=PRN.get()
    N1=Name.get()
    C1=Year.get()
    try:
        G1=gender
    except:
        messagebox.showerror("error","Select Gender!")

    D2=DOB_Name.get()
    D1=Date.get()
    M1=Mobile_Number.get()
    E1=Email.get()
    C3=Course.get()
    F1=Father_Name.get()
    M2=Mother_Name.get()
    Y1=Year.get()
    C2=City_Name.get()  

    if N1=="" or   C1=="Select Year" or D2=="" or M1== "" or E1=="" or F1=="" or M2=="" or Y1=="" or C2=="" or C3=="":
        messagebox.showerror("error","Few Data is Missing!")
    else:
        file=openpyxl.load_workbook('Student_info_data.xlsx')
        sheet=file.active
        sheet.cell(column=1,row=sheet.max_row+1,value=R1)
        sheet.cell(column=2,row=sheet.max_row,value=N1)
        sheet.cell(column=3,row=sheet.max_row,value=C1)
        sheet.cell(column=4,row=sheet.max_row,value=G1)
        sheet.cell(column=5,row=sheet.max_row,value=C3)
        sheet.cell(column=6,row=sheet.max_row,value=D1)
        sheet.cell(column=7,row=sheet.max_row,value=E1)
        sheet.cell(column=8,row=sheet.max_row,value=M1)
        sheet.cell(column=9,row=sheet.max_row,value=D2)
        sheet.cell(column=10,row=sheet.max_row,value=C2)
        sheet.cell(column=11,row=sheet.max_row,value=M2)
        sheet.cell(column=12,row=sheet.max_row,value=F1)

        file.save(r'Student_info_data.xlsx')

        try:
            img.save("c:\\Users\\Tuteshwar\\OneDrive\\Desktop\\Student registration portal image/"+str(R1)+".jpg")
        except:
            messagebox.showinfo("info","Profile picture is not available!!!")

        messagebox.showinfo("info","Successfully data entered!!!")

        Clear()

        PRN_no()

    

##############################gender#############################################
def selection():
    global gender
    value=radio.get()
    if value==1:
        gender="Male"
    else:
        gender="Female"
        

#top Frames
Label(root, text="Email : DBATU@gmail.com", width=10, height=3, bg="#f0687c", anchor='e').pack(side=TOP, fill=X)
Label(root, text="Student Registration", width=10, height=2, bg="#c36464", fg="#fff", font=('arial 20 bold')).pack(side=TOP, fill=X)

#search box to  update
Search=StringVar()
Entry(root,textvariable=Search,width=15,bd=2,font="arial 20").place(x=820,y=67)
imageicon3=PhotoImage(file="F:\\New folder (2)\\Documents\\Python\\search.png")
Srch=Button(root,text="Search",compound=LEFT,image=imageicon3,width=123,bg='#68ddfa',font="arial 13 bold")
Srch.place(x=1060,y=64)

imageicon4=PhotoImage(file="F:\\New folder (2)\\Documents\\Python\\Layer 4.png")
Update_button=Button(root,image=imageicon4,bg="#c36464")
Update_button.place(x=110,y=64)

#PRN and Date
Label(root, text="PRN no:", font='arial 13', fg=framebg ,bg=background).place(x=30,y=150)
Label(root, text="Date", font='arial 13', fg=framebg,bg=background).place(x=500,y=150)
  
PRN=IntVar()
Date = StringVar()

reg_entry = Entry(root,textvariable=PRN,width=15,font="arial 10")
reg_entry.place(x=160,y=150)

PRN_no()

today=date.today()
d1=today.strftime("%d/%m/%Y")
date_entry=Entry(root,textvariable=Date,width=15,font="arial 10")
date_entry.place(x=550,y=150)

Date.set(d1)

#Student Details
obj=LabelFrame(root,text="Student's Detail",font=20,bd=2,width=900,bg=framebg,fg=framefg,height=250,relief=GROOVE)
obj.place(x=30,y=200)

Label(obj,text="Full Name:",font="arial 13",bg=framebg,fg=framefg).place(x=30,y=50)
Label(obj,text="Course:",font="arial 13",bg=framebg,fg=framefg).place(x=30,y=100)
Label(obj,text="Gender:",font="arial 13",bg=framebg,fg=framefg).place(x=30,y=150)
Label(obj,text="Year:",font="arial 13",bg=framebg,fg=framefg).place(x=500,y=50)
Label(obj,text="Mobile Number:",font="arial 13",bg=framebg,fg=framefg).place(x=500,y=100)
Label(obj,text="Email:",font="arial 13",bg=framebg,fg=framefg).place(x=500,y=150)

Name=StringVar()
name_entry=Entry(obj,textvariable=Name,width=20,font="arial 10")
name_entry.place(x=160,y=50)

Course=StringVar()
name_entry=Entry(obj,textvariable=Course,width=20,font="arial 10")
name_entry.place(x=160,y=100)

radio=IntVar()
R1=Radiobutton(obj,text="Male",variable=radio,value=1,bg=framebg,fg=framefg,command=selection)
R1.place(x=150,y=150)

R1=Radiobutton(obj,text="Female",variable=radio,value=2,bg=framebg,fg=framefg,command=selection)
R1.place(x=200,y=150)

Year=Combobox(obj,values=['1','2','3','4'],font="Robot 10",width=17,state="r")
Year.place(x=630,y=50)
Year.set("Select Year")

Email=StringVar()
Email_entry=Entry(obj,textvariable=Email,width=20,font="arial 10")
Email_entry.place(x=630,y=100)

Mobile_Number=StringVar()
name_entry=Entry(obj,textvariable=Mobile_Number,width=20,font="arial 10")
name_entry.place(x=630,y=150)


#Personal Details
obj2=LabelFrame(root,text="Personal Detail",font=20,bd=2,width=900,bg=framebg,fg=framefg,height=220,relief=GROOVE)
obj2.place(x=30,y=470)

Label(obj2,text="Date Of Birth:",font="arial 13",bg=framebg,fg=framefg).place(x=30,y=50)
Label(obj2,text="Mother's Name:",font="arial 13",bg=framebg,fg=framefg).place(x=30,y=100)

DOB_Name=StringVar()
f_entry=Entry(obj2,textvariable=DOB_Name,width=20,font="arial 10")
f_entry.place(x=160,y=50)

Mother_Name=StringVar()
MN_entry=Entry(obj2,textvariable=Mother_Name,width=20,font="arial 10")
MN_entry.place(x=160,y=100)

Label(obj2,text="City:",font="arial 13",bg=framebg,fg=framefg).place(x=500,y=50)
Label(obj2,text="Father's Name:",font="arial 13",bg=framebg,fg=framefg).place(x=500,y=100)

City_Name=StringVar()
C_entry=Entry(obj2,textvariable=City_Name,width=20,font="arial 10")
C_entry.place(x=630,y=50)

Father_Name=StringVar()
FN_entry=Entry(obj2,textvariable=Father_Name,width=20,font="arial 10")
FN_entry.place(x=630,y=100)

#image
f=Frame(root,bd=3,bg="black",width=200,height=200,relief=GROOVE)
f.place(x=1000,y=150)

img=PhotoImage(file="F:\\New folder (2)\\Documents\\Python\\upload photo.png")
lbl=Label(f,bg="black",image=img)
lbl.place(x=0,y=0)

#Button
Button(root,text="Upload",width=19,height=2,font="arial 12 bold",bg="lightblue",command=showimage).place(x=1000,y=370)
saveButton=Button(root,text="Save",width=19,height=2,font="arial 12 bold",bg="lightgreen",command=Save)
saveButton.place(x=1000,y=450)
Button(root,text="Reset",width=19,height=2,font="arial 12 bold",bg="lightpink").place(x=1000,y=530)
Button(root,text="Exit",width=19,height=2,font="arial 12 bold",bg="grey",command=Exit).place(x=1000,y=610)









root.mainloop()